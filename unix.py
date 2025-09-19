import paramiko
#import pymysql
import openpyxl
import re
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from tqdm import tqdm

#conn_db = pymysql.connect(host = 'x.x.x.x', user='userid', password='password', db='dbname')
#curs = conn_db.cursor()

#sql = 'select ip, port, id, AES_DECRYPT(UNHEX(pw), ip) as pw, hostname from v_svr;'
#curs.execute(sql)

#result = curs.fetchall()
#conn_db.close()

# ip, port, username, password
#svr_list = [["192.168.6.53", "24477", "ian", "8282op82@#", "cyjtest", "redhat"], 
#        ["192.168.6.86", "24477", "ian", "8282op82@#", "ubuntu-test", "ubuntu"]]

svr_list = [["192.168.6.100", 22, "ian", "8282op82@#", "P2P_web", "redhat"]]
#        ["192.168.6.101", 22, "ian", "8282op82@#", "P2P_DB", "redhat"]]
ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

wb = openpyxl.Workbook()                            # 엑셀 생성
sheet = wb.active

def usec(idx, svr):
    try:
        ssh.connect(svr[0], username=svr[2], port=svr[1], password=svr[3])
        ssh.invoke_shell()
    
        stdin, stdout, stderr = ssh.exec_command("lsb_release -i")
        for line in stdout:
            Lver = line.rstrip('\n')   

        user_lst = []

        stdin, stdout, stderr = ssh.exec_command("cat /etc/passwd | awk -F: '$3 >= 1000 {print $1}' | grep -v nobody")
        for line in stdout:
            user_lst.append(line.rstrip('\n'))

        # 영문 변환 
        stdin, stdout, stderr = ssh.exec_command("LANG=en_US.UTF-8")
        lang = stdout.read().decode().strip()

        #print(svr[4])
        #print(Lver)
        thin_border = Border(left=Side(style='thin'),
                right = Side(style='thin'),
                top = Side(style='thin'),
                bottom = Side(style='thin'))

        nsheet = "sheet"+"idx"
        nsheet = wb.create_sheet(svr[4]+"("+svr[0]+")")
        nsheet.column_dimensions['B'].width = 70
        nsheet.column_dimensions['C'].width = 70
        nsheet.column_dimensions['D'].width = 85
        nsheet.cell(row=1, column=1).value = "구분코드"
        nsheet['A1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['A1'].font = openpyxl.styles.fonts.Font(bold=True)
        nsheet.cell(row=1, column=2).value = "점검리스트(cheklist)"
        nsheet['B1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['B1'].font = openpyxl.styles.fonts.Font(bold=True)
        nsheet.cell(row=1, column=3).value = "권고값"
        nsheet['C1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['C1'].font = openpyxl.styles.fonts.Font(bold=True)
        nsheet.cell(row=1, column=4).value = "서버상태"
        nsheet['D1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['D1'].font = openpyxl.styles.fonts.Font(bold=True)
        nsheet.cell(row=1, column=5).value = "보안수준"
        nsheet['E1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['E1'].font = openpyxl.styles.fonts.Font(bold=True)
# ================================================== AC-01a ==================================================
    
        AC01A = ["AC01-A", "패스워드 복잡도 설정이 적용되었는가?", "lcredit : -1\n ocredit : -1\n dcredit : -1", "", "양호"]
        ac01a = []

        stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/pam.d/common-password | grep -v '#' | grep lcredit | awk -F'lcredit' '{print $2}' | cut -c 2-3")
        ac01a.append(stdout.read().decode().strip())  

        stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/pam.d/common-password | grep -v '#' | grep dcredit | awk -F'dcredit' '{print $2}' | cut -c 2-3")
        ac01a.append(stdout.read().decode().strip())   

        stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/pam.d/common-password | grep -v '#' | grep ocredit | awk -F'ocredit' '{print $2}' | cut -c 2-3")
        ac01a.append(stdout.read().decode().strip())   

        AC01A[3] = "lcredit : "+ac01a[0]+"\n"+"dcredit : "+ac01a[1]+"\n"+"ocredit : "+ac01a[2]
        for i in ac01a:
            if len(i) == 0:
                AC01A[4] = "취약"

        nsheet.append(AC01A)
        if AC01A[4] == "취약":
            nsheet['E2'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E2'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-01b ==================================================
        AC01B = ["AC01-B", "패스워드의 최소 암호길이가 설정 되었는가?", "minlen : 8", "", ""]

        ac01b = 0

        stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/pam.d/common-password | grep -v '#' | grep minlen | awk -F'minlen' '{print $2}' | cut -c 2-3")
        ac01b = stdout.read().decode().strip()

        AC01B[3] = "minlen : "+ac01b

        AC01B[4] = "취약" if not ac01b or int(ac01b) < 8 else "양호"

        nsheet.append(AC01B)
        if AC01B[4] == "취약":
            nsheet['E3'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E3'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-02a ==================================================
        AC02A = ["AC02-A", "패스워드 최근 암호 기억이 설정되었는가?", "remember : 2", "", ""]

        ac02a = 0

        stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/pam.d/common-password | grep -v '#' | grep remember | awk -F'remember' '{print $2}' | cut -c 2-3")

        ac02a = stdout.read().decode().strip()

        AC02A[3] = "remember : "+ac02a

        AC02A[4] = "취약" if not ac02a or int(ac02a) < 2 else "양호"

        nsheet.append(AC02A)
        if AC02A[4] == "취약":
            nsheet['E4'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E4'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-02b ==================================================X
        AC02B = ["AC02-B", "패스워드 최대 사용기간이 설정되었는가?", "[PASS_MAX_DAYS] : 90", "", ""]

        ac02b = []

        stdin, stdout, stderr = ssh.exec_command("cat /etc/login.defs | grep -v '#' | grep PASS_MAX_DAYS | awk '{print $2}'")

        ac02b.append(stdout.read().decode().strip())
        
        AC02B[4] = "취약" if not ac02b[0] or int(ac02b[0]) > 90 else "양호"

        for user in user_lst:
            stdin, stdout, stderr = ssh.exec_command("LANG=en_US.UTF-8 && sudo chage -l "+user+" | grep Maximum | awk -F: '{print $2}'")
            ret = stdout.read().decode().strip()
            if int(ret) > 90:
                AC02B[4] = "취약"
                ac02b.append(user+"="+ret)
            else:
                ac02b.append(ret)

        AC02B[3] = "PASS_MAX_DAYS : "+ac02b[0]+", 각 사용자 패드워드 설정 : "+", ".join(ac02b[1:])

        nsheet.append(AC02B)
        if AC02B[4] == "취약":
            nsheet['E5'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E5'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-02c ==================================================
        AC02C = ["AC02-C", "패스워드 최소 사용기간이 설정되었는가?", "[PASS_MIN_DAYS] : 7", "", ""]

        ac02c = 0

        stdin, stdout, stderr = ssh.exec_command("cat /etc/login.defs | grep -v '#' | grep PASS_MIN_DAYS | awk '{print $2}'")

        ac02c = stdout.read().decode().strip()

        AC02C[3] = "PASS_MIN_DAYS : "+ac02c

        AC02C[4] = "취약" if not ac02c or int(ac02c) < 7 else "양호"

        nsheet.append(AC02C)
        if AC02C[4] == "취약":
            nsheet['E6'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E6'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-03 ==================================================
        AC03 = ["AC03", "계정 잠금 임계값이 설정되었는가?", "[deny] : 5 [unlock_time] : 1800", "", ""]
        
        ac03 = []

        stdin, stdout, stderr = ssh.exec_command("cat /etc/pam.d/common-auth | grep auth | grep default=die | grep -E 'pam_faillock.so|pam_tally2.so' | grep -o 'deny=[0-9]*' | awk -F= '{print $2}'")
        ac03.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("cat /etc/pam.d/common-auth | grep auth | grep default=die | grep -E 'pam_faillock.so|pam_tally2.so' | grep -o 'unlock_time=[0-9]*' | awk -F= '{print $2}'")
        ac03.append(stdout.read().decode().strip())

        if ac03[0] == "" or int(ac03[0]) > 5:
            AC03[4] = "취약"
            AC03[3] = "deny : "+ac03[0]+", unlock_time : "+ac03[1]
        elif int(ac03[0]) > 5 or int(ac03[1]) > 1800:
            AC03[4] = "취약"
            AC03[3] = "deny : "+ac03[0]+", unlock_time : "+ac03[1]
        else:
            AC03[4] = "양호"
            AC03[3] = "deny : "+ac03[0]+", unlock_time : "+ac03[1]

        nsheet.append(AC03)
        if AC03[4] == "취약":
            nsheet['E7'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E7'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04A ==================================================
        AC04A = ["AC04-A", "사용자 계정, 그룹 리스트 정보 파일에 대한 접근권한이 제한되어 있는가?", "passwd : -rw-r--r--. 1 root root\ngroup : -rw-r--r--. 1 root root", "", ""]

        ac04a = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/passwd")
        ac04a.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/passwd")
        ac04a.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/group")
        ac04a.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/group")
        ac04a.append(stdout.read().decode().strip())

        AC04A[3] = "/etc/passwd : "+ac04a[0]+" "+ac04a[1]+", "+"/etc/group : "+ac04a[2]+" "+ac04a[3]

        AC04A[4] = "취약" if int(ac04a[0]) > 644 or ac04a[1] != "root" or int(ac04a[2]) > 644 or ac04a[3] != "root" else "양호"

        nsheet.append(AC04A)
        if AC04A[4] == "취약":
            nsheet['E8'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E8'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04B ==================================================
        AC04B = ["AC04-B", "사용자 계정 암호 파일에 대한 접근 권한이 제한되어 있는가?", "shadow : -r--------. 1 root root", "", ""]

        ac04b = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/shadow")
        ac04b.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/shadow")
        ac04b.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("sudo passwd -S ian | awk '{print $2}'")
        ac04b.append(stdout.read().decode().strip())

        AC04B[3] = "/etc/shadow : "+ac04b[0]+" "+ac04b[1]

        AC04B[4] = "취약" if int(ac04b[0]) > 400 or ac04b[1] != "root" or 'P' not in ac04b[2] else "양호"

        nsheet.append(AC04B)
        if AC04B[4] == "취약":
            nsheet['E9'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E9'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04C ==================================================
        AC04C = ["AC04-C", "/etc/hosts 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/hosts : -rw-r--r--. 1 root root", "", ""]

        ac04c = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/hosts")
        ac04c.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/hosts")
        ac04c.append(stdout.read().decode().strip())

        AC04C[3] = "/etc/hosts : "+ac04c[0]+" "+ac04c[1]

        AC04C[4] = "취약" if int(ac04c[0]) > 644 or ac04c[1] != "root" else "양호"

        nsheet.append(AC04C)
        if AC04C[4] == "취약":
            nsheet['E10'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E10'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04D ==================================================
        AC04D = ["AC04-D", "/etc/services 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/services : -rw-r--r--. 1 root root", "", ""]

        ac04d = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/services")
        ac04d.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/services")
        ac04d.append(stdout.read().decode().strip())

        AC04D[3] = "/etc/services : "+ac04d[0]+" "+ac04d[1]

        AC04D[4] = "취약" if int(ac04d[0]) > 644 or ac04d[1] != "root" else "양호"

        nsheet.append(AC04D)
        if AC04D[4] == "취약":
            nsheet['E11'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E11'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04E ==================================================
        AC04E = ["AC04-E", "일반 사용자가 중요 명령어를 실행하지 못하도록 설정되었는가?", "/usr/bin/last : -rw-------. 1 root root", "", ""]

        ac04e = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /usr/bin/last")
        ac04e.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /usr/bin/last")
        ac04e.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("command -v ifconfig")
        ac04e.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /usr/sbin/ifconfig")
        ac04e.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /usr/sbin/ifconfig")
        ac04e.append(stdout.read().decode().strip())

        if not ac04e[2]:
            AC04E[3] = "/usr/bin/last : "+ac04e[0]+" "+ac04e[1]
            AC04E[4] = "취약" if int(ac04e[0]) > 700 or ac04e[1] != "root" else "양호"
        else:
            AC04E[3] = "/usr/bin/last : "+ac04e[0]+" "+ac04e[1]+"\n"+"/usr/sbin/ifconfig : "+ac04e[3]+" "+ac04e[4]+"\n"
            AC04E[4] = "취약" if int(ac04e[0]) > 700 or ac04e[1] != "root" or int(ac04e[3]) > 700 or ac04e[4] != "root" else "양호"

        nsheet.append(AC04E)
        if AC04E[4] == "취약":
            nsheet['E12'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E12'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04F ==================================================
        AC04F = ["AC04-F", "사용자 환경 설정 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/profile : -rwxr-xr-x. 1 root root", "", ""]

        ac04f = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/profile")
        ac04f.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/profile")
        ac04f.append(stdout.read().decode().strip())

        AC04F[3] = "/etc/profile : "+ac04f[0]+" "+ac04f[1]+"\n"

        AC04F[4] = "취약" if int(ac04f[0]) > 755 or ac04f[1] != "root" else "양호"

        nsheet.append(AC04F)
        if AC04F[4] == "취약":
            nsheet['E13'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E13'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04G ==================================================
        AC04G = ["AC04-G", "네트워크 서비스 설정 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/xinetd.conf : -rw-------. 1 root root", "", ""]

        ac04g = []

        stdin, stdout, stderr = ssh.exec_command("ls -al /etc/xinetd.conf")
        ac04g.append(stdout.read().decode().strip())

        if len(ac04g[0]) == 0:
            AC04G[3] = "/etc/xinetd.conf 파일 없음"
            AC04G[4] = "양호"
        else:
            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/xinetd.conf")
            ac04g.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/xinetd.conf")
            ac04g.append(stdout.read().decode().strip())

            AC04G[3] = "/etc/xinetd.conf : "+ac04f[0]+" "+ac04f[1]+"\n"

            AC04G[4] = "취약" if int(ac04g[0]) > 600 or ac04g[1] != "root" else "양호"

        nsheet.append(AC04G)
        if AC04G[4] == "취약":
            nsheet['E14'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E14'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04H ==================================================
        AC04H = ["AC04-H", "FTP 접근제어 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/vsftpd.ftpusers, /etc/vsftpd.user_list : -rw-------. 1 root root", "", ""]

        ac04h = []

        stdin, stdout, stderr = ssh.exec_command("sudo systemctl status vsftpd | grep Active: | awk -F' ' '{print $2}'")
        ac04h.append(stdout.read().decode().strip())

        if ac04h[0] == "" or ac04h[0] == "inactive":
            AC04H[3] = "vsftp 서비스 기동되어있지 않음"
            AC04H[4] = "양호"
        else:
            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/vsftpd/ftpusers")
            ac04h.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/vsftpd/ftpusers")
            ac04h.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/vsftpd/user_list")
            ac04h.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/vsftpd/user_list")
            ac04h.append(stdout.read().decode().strip())

            AC04H[3] = "/etc/vsftpd/ftpusers : "+ac04h[1]+" "+ac04h[2]+", "+"/etc/vsftpd/user_list : "+ac04h[3]+" "+ac04h[4]
            AC04H[4] = "취약" if int(ac04h[1]) > 600 or ac04h[2] != "root" or int(ac04h[3]) > 600 or ac04h[4] != "root" else "양호"

        nsheet.append(AC04H)
        if AC04H[4] == "취약":
            nsheet['E15'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E15'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04I ==================================================
        AC04I = ["AC04-I", "NFS 설정 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/exports : -rw-------. 1 root root", "", ""]

        ac04i = []

        stdin, stdout, stderr = ssh.exec_command("ls -al /etc/exports")
        ac04i.append(stdout.read().decode().strip())

        if len(ac04i[0]) == 0:
            AC04I[3] = "NFS 데몬 설치 되어있지 않음"
            AC04I[4] = "양호"
        else:
            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/exports")
            ac04i.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%A' /etc/exports")
            ac04i.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/exports")
            ac04i.append(stdout.read().decode().strip())

            AC04I[3] = "/etc/exports : "+ac04i[1]+" "+ac04i[3]
            AC04I[4] = "취약" if "w" in ac04i[1][4:] or ac04i[3] != "root" else "양호"

        nsheet.append(AC04I)
        if AC04I[4] == "취약":
            nsheet['E16'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E16'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04J ==================================================
        AC04J = ["AC04-J", "패스워드 규칙 설정 파이에 대한 접근 권한이 제한되어 있는가?", "/etc/pam.d/system-auth, password-auth : 타사용자 쓰기권한 없음, 소유자 root", "", ""]

        ac04j = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/pam.d/common-auth")
        ac04j.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%A' /etc/pam.d/common-auth")
        ac04j.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/pam.d/common-auth")
        ac04j.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/pam.d/common-auth")
        ac04j.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%A' /etc/pam.d/common-auth")
        ac04j.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/pam.d/common-auth")
        ac04j.append(stdout.read().decode().strip())

        AC04J[3] = "/etc/pam.d/common-auth : "+ac04j[0]+" "+ac04j[2]+", "+"/etc/pam.d/common-auth : "+ac04j[3]+" "+ac04j[5]
        AC04J[4] = "취약" if "w" in ac04j[1][4:] or ac04j[2] != "root" or "w" in ac04j[4][4:] or ac04j[5] != "root" else "양호"

        nsheet.append(AC04J)
        if AC04J[4] == "취약":
            nsheet['E17'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E17'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04K ==================================================
        AC04K = ["AC04-K", "cron 설정 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/cron.allow, cron.deny : 타사용자 쓰기권한 없음, 소유자 root", "", ""]

        ac04k = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/cron.allow")
        ac04k.append(stdout.read().decode().strip())

        if len(ac04k[0]) != 0:
            stdin, stdout, stderr = ssh.exec_command("stat -c '%A' /etc/cron.allow")
            ac04k.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/cron.allow")
            ac04k.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/cron.deny")
            ac04k.append(stdout.read().decode().strip())
            if len(ac04k[3]) != 0:
                stdin, stdout, stderr = ssh.exec_command("stat -c '%A' /etc/cron.deny")
                ac04k.append(stdout.read().decode().strip())

                stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/cron.deny")
                ac04k.append(stdout.read().decode().strip())

                AC04K[3] = "/etc/cron.allow : "+ac04k[0]+" "+ac04k[2]+", "+"/etc/cron.deny : "+ac04k[3]+" "+ac04k[5]
                AC04K[4] = "취약" if "w" in ac04k[1][4:] or ac04k[2] != "root" or "w" in ac04k[4][4:] or ac04k[5] != "root" else "양호"
            else:
                AC04K[3] = "/etc/cron.allow : "+ac04k[0]+" "+ac04k[2]+" /etc/cron.deny 파일 없음"
                AC04K[4] = "취약" if "w" in ac04k[1][4:] or ac04k[2] != "root" else "양호"
        else:
            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/cron.deny")
            ac04k.append(stdout.read().decode().strip())
            if len(ac04k[1]) != 0:
                stdin, stdout, stderr = ssh.exec_command("stat -c '%A' /etc/cron.deny")
                ac04k.append(stdout.read().decode().strip())

                stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/cron.deny")
                ac04k.append(stdout.read().decode().strip())

                AC04K[3] = "/etc/cron.deny : "+ac04k[1]+" "+ac04k[3]+" /etc/cron.allow 파일 없음"
                AC04K[4] = "취약" if "w" in ac04k[2][4:] or ac04k[3] != "root" else "양호"
            else:
                AC04K[3] = "/etc/cron.deny, cron.allow 파일 없음"
                AC04K[4] = "양호"

        nsheet.append(AC04K)
        if AC04K[4] == "취약":
            nsheet['E18'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E18'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04L ==================================================
        AC04L = ["AC04-L", "주요 백업 파일 접근 권한이 제한되어 있는가?", "passwd, services, hosts, wtmp, btmp, auth.log 백업 파일 유무 및 권한 확인", "", ""]

        stdin, stdout, stderr = ssh.exec_command("ls -al /etc/passwd* /etc/services* /etc/hosts* /var/log/wtmp* /var/log/btmp* /var/log/auth.log* | awk -F' ' '{print $9}'")
        file_lst = []
        for line in stdout:
            file_name = line.rstrip('\n')
            if file_name not in ["/etc/passwd", "/etc/services", "/etc/hosts", "/var/log/wtmp", "/var/log/btmp", "/var/log/secure", "/etc/hosts.allow", "/etc/hosts.deny"]:
                file_lst.append(file_name)

        AC04L[4] = "양호"
        for file in file_lst:
            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' "+file)
            perm = stdout.read().decode().strip()

            stdin, stdout, stderr = ssh.exec_command("stat -c '%U' "+file)
            own = stdout.read().decode().strip()
            if file == "/etc/passwd-":
                if int(perm) > 644 or own != "root":
                    AC04L[4] = "취약"
            else:
                if int(perm) != 600 or own != "root":
                    AC04L[4] = "취약"
            AC04L[3] += file+" : "+perm+" "+own+" "

        nsheet.append(AC04L)
        if AC04L[4] == "취약":
            nsheet['E19'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E19'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-05A ==================================================
        AC05A = ["AC05-A", "root 계정의 UMASK 설정은 적절한가?", "UMASK : 022, 027", "", ""]

        ac05a = []

        stdin, stdout, stderr = ssh.exec_command("sudo su - root -c 'umask'")
        ac05a.append(stdout.read().decode().rstrip('\n'))

        AC05A[3] = "UMASK : "+ac05a[0]
        AC05A[4] = "취약" if int(ac05a[0]) > 27 else "양호"

        nsheet.append(AC05A)
        if AC05A[4] == "취약":
            nsheet['E20'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E20'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-05B ==================================================
        AC05B = ["AC05-B", "일반 계정의 UMASK 설정은 적절한가?", "UMASK : 022, 027", "", ""]

        ac05b = []

        AC05B[4] = "양호"
        for user in user_lst:
            stdin, stdout, stderr = ssh.exec_command("sudo su - "+user+" -c 'umask'")
            ret = int(stdout.read().decode().strip())
            if ret > 27:
                ac05b.append(user+" : "+str(ret))
                AC05B[4] = "취약"
            else:
                ac05b.append(user+" : "+str(ret))

        AC05B[3] = ", ".join(ac05b)

        nsheet.append(AC05B)
        if AC05B[4] == "취약":
            nsheet['E21'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E21'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-05C ==================================================
        AC05C = ["AC05-C", "root 계정의 원격접속은 제한되어 있는가?", "/etc/ssh/sshd_config : PermitRootLogin no", "", ""]

        ac05c = []

        stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/ssh/sshd_config | grep 'PermitRootLogin' | grep -v '#' | awk '{print $2}'")
        ac05c.append(stdout.read().decode().strip())
        
        AC05C[3] = "PermitRootLogin : "+ac05c[0]
        AC05C[4] = "양호" if ac05c[0] == "no" else "취약"

        nsheet.append(AC05C)
        if AC05C[4] == "취약":
            nsheet['E22'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E22'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-06 ==================================================
        AC06 = ["AC06", "Session Timeout을 적용하고 있는가?", "TMOUT : 1800", "", ""]

        ac06 = []

        stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/profile | grep TMOUT | grep -v '#' | awk -F= '{print $2}'")
        ac06.append(stdout.read().decode().strip())

        AC06[4] = "양호"
        if len(ac06[0]) == 0:
            AC06[4] = "취약"
            ac06[0] = "TMOUT 설정 없음"
            AC06[3] = "/etc/profile TMOUT : "+ac06[0]
        elif int(ac06[0]) > 1800:
            AC06[4] = "취약"
            AC06[3] = "/etc/profile TMOUT : "+ac06[0]
        else:
            for user in user_lst:
                stdin, stdout, stderr = ssh.exec_command("cat /home/"+user+"/.bash_profile | grep TMOUT | grep -v '#' | awk -F= '{print $2}'")
                ret = stdout.read().decode().strip()
                if len(ret) == 0:
                    ac06.append(user+" : 0")
                elif int(ret) > 1800:
                    AC06[4] = "취약"
                    ac06.append(user+" TMOUT : "+ret)
                else:
                    u01.append(user+" TMOUT : "+ret)

            AC06[3] = "/etc/profile TMOUT : "+ac06[0]+", .bash_profile TMOUT : ".join(ac06[1:])

        nsheet.append(AC06)
        if AC06[4] == "취약":
            nsheet['E23'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E23'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== SS-01 ==================================================
        SS01 = ["SS01", "NFS 사용 시 안전한 설정을 적용 하였는가?", "공유 디렉토리 : root, 사용자 홈디렉토리 공유 금지, 권한 : everyone 공유 금지", "", ""]

        ss01 = []

        stdin, stdout, stderr = ssh.exec_command("sudo systemctl status nfs-server | grep Active: | awk -F' ' '{print $2}'")
        ss01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("sudo netstat -antp | grep 2049 | grep ESTABLISHED | wc -l")
        ss01.append(stdout.read().decode().strip())

        if ss01[0] == "Active" and int(ss01[1]) > 0:
            stdin, stdout, stderr = ssh.exec_command("cat /etc/exports")
            ss01.append(stdout.read().decode().strip())
            if "/home" in sso1[2] or "/root" in ss01[2] or "*" in ss01[2]:
                SS01[4] = "취약"
                SS01[3] = "NFS-Server 서비스: "+ss01[0]+", NFS 연결 수: "+ss01[1]+", /etc/exports 설정: "+ss01[2]
            else:
                SS01[4] = "양호"
                SS01[3] = "NFS-Server 서비스: "+ss01[0]+", NFS 연결 수: "+ss01[1]+", /etc/exports 설정: "+ss01[2]
        elif ss01[0] == "Active":
            stdin, stdout, stderr = ssh.exec_command("cat /etc/exports")
            ss01.append(stdout.read().decode().strip())
            if "/home" in u01[2] or "/root" in u01[2] or "*" in u01[2]:
                SS01[4] = "취약"
                SS01[3] = "NFS-Server 서비스: "+ss01[0]+", NFS 연결 수: "+ss01[1]+", /etc/exports 설정: "+ss01[2]
            else:
                SS01[4] = "양호"
                SS01[3] = "NFS-Server 서비스: "+ss01[0]+", NFS 연결 수: "+ss01[1]+", /etc/exports 설정: "+ss01[2]
        elif int(ss01[1]) > 0:
            stdin, stdout, stderr = ssh.exec_command("cat /etc/fstab | grep nfs")
            ss01.append(stdout.read().decode().strip())
            if "/root" in ss01[2] or "/home" in ss01[2]:
                SS01[4] = "취약"
                SS01[3] = "NFS Client 동작중, NFS 연결 수: "+ss01[1]+", /etc/fstab 설정: "+ss01[2]
            else:
                SS01[4] = "양호"
                SS01[3] = "NFS Client 동작중, NFS 연결 수: "+ss01[1]+", /etc/fstab 설정: "+ss01[2]
        else:
            SS01[4] = "양호"
            SS01[3] = "NFS Server 및 Client로 동작안함"

        nsheet.append(SS01)
        if SS01[4] == "취약":
            nsheet['E24'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E24'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== SS-02A ==================================================
        SS02A = ["SS02-A", "원격에서 인증없이 접속하여 명령 실행이 가능한 서비스 사용을 중지하였는가?", "r계열 서비스 중지 확인", "", ""]

        ss02a = []

        stdin, stdout, stderr = ssh.exec_command("rpm -qa | grep -E 'rsh|rlogin|rexec' | wc -l")
        ss02a.append(stdout.read().decode().strip())
        
        SS02A[3] = "r계열 서비스 설치수: "+ss02a[0]
        SS02A[4] = "양호" if int(ss02a[0]) == 0 else "취약"

        nsheet.append(SS02A)
        if SS02A[4] == "취약":
            nsheet['E25'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E25'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== SS-02B ==================================================
        SS02B = ["SS02-B", "불필요한 서비스는 중지 또는 제거하였는가?", "불필요 서비스 중지 확인", "", ""]

        ss02b = []
        stdin, stdout, stderr = ssh.exec_command("sudo systemctl status time")
        ss02b.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("sudo systemctl list-units --type=service | grep -E 'bootp|chargen|cmsd|daytime|discard|echo|finger|netstat|rusersd|sprayed|systat|tftp|ttdbserverd|uucp'")
        for line in stdout:
            ss02b.append(line.rstrip('\n'))

        if len(ss02b) <= 1 and len(ss02b[0]) == 0:
            SS02B[4] = "양호"
            SS02B[3] = "불필요한 서비스 없음"
        else:
            SS02B[4] = "취약"
            SS02B[3] = "불필요 서비스 리스트 : "+", ".join(ss02b)

        nsheet.append(SS02B)
        if SS02B[4] == "취약":
            nsheet['E26'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E26'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== SS-03A ==================================================
        SS03A = ["SS03-A", "SNMP Default Community String이 변경되었는가?", "Community string : public, private 등 default 값 사용 금지 ", "", ""]

        ss03a = []

        stdin, stdout, stderr = ssh.exec_command("sudo systemctl status snmpd | grep Active: | awk -F' ' '{print $2}'")
        ss03a.append(stdout.read().decode().strip())

        if ss03a[0] == "inactive":
            SS03A[4] = "양호"
            SS03A[3] = "SNMP 서비스 동작 안함"
        else:
            stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/snmp/snmpd.conf | grep -v '#' | grep com2sec")
            ss03a.append(stdout.read().decode().strip())
            if "public" in ss03a[1]:
                SS03A[4] = "취약"
                SS03A[3] = "SNMP 서비스 : "+ss03a[0]+", community 설정 : "+ss03a[1]
            else:
                SS03A[4] = "양호"
                SS03A[3] = "SNMP 서비스 : "+ss03a[0]+", community 설정 : "+ss03a[1]

        nsheet.append(SS03A)
        if SS03A[4] == "취약":
            nsheet['E27'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E27'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== SS-03B ==================================================
        SS03B = ["SS03-B", "Anonymouse FTP가 중지되어 있는가?", "/etc/passwd, shadow : ftp, anonymous 계정 없음 확인, anonymous_enable : No", "", ""]

        ss03b = []

        stdin, stdout, stderr = ssh.exec_command("sudo systemctl status vsftpd | grep Active: | awk -F' ' '{print $2}'")
        ss03b.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/passwd | grep -E 'ftp|anonymous' | awk -F: '{print $1}'")
        for line in stdout:
            ss03b.append(line.rstrip('\n'))

        stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/shadow | grep -E 'ftp|anonymous' | awk -F: '{print $1}'")
        for line in stdout:
            ss03b.append(line.rstrip('\n'))

        if ss03b[0] == "Active":
            stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/vsftpd/vsftpd.conf | grep anonymous | grep -v '#' | awk -F= '{print $2}'")
            ret = stdout.read().decode().strip()
            if ret == "YES" or "ftp" in ss03b or "anonymous" in ss03b:
                SS03B[4] = "취약"
                if len(ss03b[1:]) == 0:
                    SS03B[3] = "passwd, shadow에 ftp, anonymous계정 없음, anonymous_enable : "+ret
                else:
                    SS03B[3] = "passwd, shadow : "+", ".join(ss03b[1:])+", anonymous_enable : "+ret
            else:
                SS03B[4] = "양호"
                SS03B[3] = "passwd, shadow에 ftp, anonymous계정 없음, anonymous_enable : "+ret
        else:
            if "ftp" in ss03b or "anonymous" in ss03b:
                SS03B[4] = "취약"
                SS03B[3] = "passwd, shadow : "+", ".join(ss03b[1:])+", vsftpd 서비스 없음"
            else:
                SS03B[4] = "양호"
                SS03B[3] = "passwd, shadow에 ftp, anonymous계정 없음, vsftpd 서비스 없음"

        nsheet.append(SS03B)
        if SS03B[4] == "취약":
            nsheet['E28'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E28'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== SS-04 ==================================================
        SS04 = ["SS04", "root 계정의 Path 설정이 안전하게 적용되어 있는가?", "Path 환경 변수 : 현재 디렉토리(.) 미포함", "", ""]

        ss04 = []

        stdin, stdout, stderr = ssh.exec_command("sudo su - root -c 'env | grep PATH'")
        ss04.append(stdout.read().decode().strip())

        SS04[3] = "PATH 환경 변수 : "+ss04[0]
        SS04[4] = "취약" if ".:" in ss04[0] else "양호"

        nsheet.append(SS04)
        if SS04[4] == "취약":
            nsheet['E29'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E29'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== LP-01 ==================================================
        LP01 = ["LP01", "서버 접속 로그에 대한 접근 권한을 제한 하는가?", "/var/log/wtmp, btmp, secure : root 600, messages : root 644", "", ""]

        lp01 = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /var/log/wtmp")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /var/log/wtmp")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /var/log/btmp")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /var/log/btmp")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /var/log/secure")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /var/log/secure")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /var/log/messages")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /var/log/messages")
        lp01.append(stdout.read().decode().strip())

        LP01[3] = "/var/log/wtmp, btmp, secure : "+lp01[0]+" "+lp01[1]+", "+lp01[2]+" "+lp01[3]+", "+lp01[3]+" "+lp01[4]+"/var/log/messages : "+lp01[5]+" "+lp01[6]
        LP01[4] = "취약" if lp01[0] != "600" or lp01[1] != "root" or lp01[2] != "600" or lp01[3] != "root" or lp01[4] != "600" or lp01[5] != "root" or lp01[6] != "644" or lp01[7] != "root" else "양호"

        nsheet.append(LP01)
        if LP01[4] == "취약":
            nsheet['E30'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E30'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# =================================================== END ==================================================

        for i in range(1, 6):
            for j in range(1, 31):
                nsheet.cell(row=j, column=i).border = thin_border

    except paramiko.ssh_exception.NoValidConnectionsError as 접속불가:
        print("\n["+svr[4]+"]"+" [접속불가] IP 및 PORT 확인")
    
    except paramiko.ssh_exception.AuthenticationException as 접속불가:
        print("\n["+svr[4]+"]"+" [접속불가] 사용자 이름 또는 패스워드 확인")

    except paramiko.ssh_exception.PasswordRequiredException as 접속불가:
        print("\n["+svr[4]+"]"+" [접속불가] Password Required")

    except paramiko.ssh_exception.SSHException as 접속불가:
        print("\n["+svr[4]+"]"+" [접속불가] SSHException")
    finally:
        ssh.close()

def rsec(idx, svr):
    try:
        #ssh.connect(svr[0], username=svr[2], port=svr[1], password=svr[3].decode('utf-8'))
        ssh.connect(svr[0], username=svr[2], port=svr[1], password=svr[3])
        ssh.invoke_shell()
    
        stdin, stdout, stderr = ssh.exec_command("cat -n /etc/system-release | cut -d'.' -f1 | awk '{print $NF}'")
        for line in stdout:
            Lver = line.rstrip('\n')

        user_lst = []

        stdin, stdout, stderr = ssh.exec_command("cat /etc/passwd | awk -F: '$3 >= 1000 {print $1}' | grep -v nobody")
        for line in stdout:
            user_lst.append(line.rstrip('\n'))
        
        # 영문 변환 
        stdin, stdout, stderr = ssh.exec_command("LANG=en_US.UTF-8")
        lang = stdout.read().decode().strip()

        #print(svr[4])
        thin_border = Border(left=Side(style='thin'),
                right = Side(style='thin'),
                top = Side(style='thin'),
                bottom = Side(style='thin'))

        nsheet = "sheet"+"idx"
        nsheet = wb.create_sheet(svr[4]+"("+svr[0]+")")
        nsheet.column_dimensions['B'].width = 70
        nsheet.column_dimensions['C'].width = 70
        nsheet.column_dimensions['D'].width = 85
        nsheet.cell(row=1, column=1).value = "구분코드"
        nsheet['A1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['A1'].font = openpyxl.styles.fonts.Font(bold=True)
        nsheet.cell(row=1, column=2).value = "점검리스트(cheklist)"
        nsheet['B1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['B1'].font = openpyxl.styles.fonts.Font(bold=True)
        nsheet.cell(row=1, column=3).value = "권고값"
        nsheet['C1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['C1'].font = openpyxl.styles.fonts.Font(bold=True)
        nsheet.cell(row=1, column=4).value = "서버상태"
        nsheet['D1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['D1'].font = openpyxl.styles.fonts.Font(bold=True)
        nsheet.cell(row=1, column=5).value = "보안수준"
        nsheet['E1'].fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        nsheet['E1'].font = openpyxl.styles.fonts.Font(bold=True)

# ================================================== AC-01a ==================================================
    
        AC01A = ["AC01-A", "패스워드 복잡도 설정이 적용되었는가?", "lcredit : -1\n ocredit : -1\n dcredit : -1", "", "양호"]
        ac01a = []

        stdin, stdout, stderr = ssh.exec_command("cat /etc/security/pwquality.conf | grep -v '#' | grep lcredit | awk '{print $3}'")
        ac01a.append(stdout.read().decode().strip())


        stdin, stdout, stderr = ssh.exec_command("cat /etc/security/pwquality.conf | grep -v '#' | grep ocredit | awk '{print $3}'")
        ac01a.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("cat /etc/security/pwquality.conf | grep -v '#' | grep dcredit | awk '{print $3}'")
        ac01a.append(stdout.read().decode().strip())

        AC01A[3] = "lcredit : "+ac01a[0]+"\n"+"ocredit : "+ac01a[1]+"\n"+"dcredit : "+ac01a[2]
        for i in ac01a:
            if len(i) == 0:
                AC01A[4] = "취약"

        nsheet.append(AC01A)
        if AC01A[4] == "취약":
            nsheet['E2'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E2'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-01b ==================================================
        AC01B = ["AC01-B", "패스워드의 최소 암호길이가 설정 되었는가?", "minlen : 8", "", ""]

        ac01b = 0

        stdin, stdout, stderr = ssh.exec_command("cat /etc/security/pwquality.conf | grep -v '#' | grep minlen | awk '{print $3}'")

        ac01b = stdout.read().decode().strip()

        AC01B[3] = "minlen : "+ac01b

        AC01B[4] = "취약" if not ac01b or int(ac01b) < 8 else "양호"

        nsheet.append(AC01B)
        if AC01B[4] == "취약":
            nsheet['E3'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E3'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-02a ==================================================
        AC02A = ["AC02-A", "패스워드 최근 암호 기억이 설정되었는가?", "remember : 2", "", ""]

        ac02a = 0

        stdin, stdout, stderr = ssh.exec_command("cat /etc/pam.d/password-auth | grep -v '#' | grep password | grep requisite | grep pam_pwquality.so| grep remember | awk -F'remember' '{print $2}' | cut -c 2-3")

        ac02a = stdout.read().decode().strip()

        AC02A[3] = "remember : "+ac02a

        AC02A[4] = "취약" if not ac02a or int(ac02a) < 2 else "양호"

        nsheet.append(AC02A)
        if AC02A[4] == "취약":
            nsheet['E4'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E4'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-02b ==================================================
        AC02B = ["AC02-B", "패스워드 최대 사용기간이 설정되었는가?", "[PASS_MAX_DAYS] : 90", "", ""]

        ac02b = []

        stdin, stdout, stderr = ssh.exec_command("cat /etc/login.defs | grep -v '#' | grep PASS_MAX_DAYS | awk '{print $2}'")

        ac02b.append(stdout.read().decode().strip())
        
        AC02B[4] = "취약" if not ac02b[0] or int(ac02b[0]) > 90 else "양호"

        for user in user_lst:
            stdin, stdout, stderr = ssh.exec_command("LANG=en_US.UTF-8 && sudo chage -l "+user+" | grep Maximum | awk -F: '{print $2}'")
            ret = stdout.read().decode().strip()
            err = stderr.read().decode().strip()
            '''
            if int(ret) > 90:
                AC02B[4] = "취약"
                ac02b.append(user+"="+ret)
            else:
                ac02b.append(ret)
            '''
            if not ret:
                ac02b.append(user + " = 값 없음 (실패 또는 권한 문제)")
                continue
            try:
                days = int(ret)
                if days > 90:
                    AC02B[4] = "취약"
                    ac02b.append(user+"="+ret)
                else:
                    ac02b.append(user+"="+ret)
            except ValueError:
                ac02b.append(user+"=숫자 아님 ("+ret+")")


        AC02B[3] = "PASS_MAX_DAYS : "+ac02b[0]+", 각 사용자 패스워드 설정 : "+", ".join(ac02b[1:])

        nsheet.append(AC02B)
        if AC02B[4] == "취약":
            nsheet['E5'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E5'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-02c ==================================================
        AC02C = ["AC02-C", "패스워드 최소 사용기간이 설정되었는가?", "[PASS_MIN_DAYS] : 7", "", ""]

        ac02c = 0

        stdin, stdout, stderr = ssh.exec_command("cat /etc/login.defs | grep -v '#' | grep PASS_MIN_DAYS | awk '{print $2}'")

        ac02c = stdout.read().decode().strip()

        AC02C[3] = "PASS_MIN_DAYS : "+ac02c

        AC02C[4] = "취약" if not ac02c or int(ac02c) < 7 else "양호"

        nsheet.append(AC02C)
        if AC02C[4] == "취약":
            nsheet['E6'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E6'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-03 ==================================================
        AC03 = ["AC03", "계정 잠금 임계값이 설정되었는가?", "[deny] : 5\n[unlock_time] : 1800\n[no_magic_root  reset] : 존재", "", ""]
        
        ac03 = []

        stdin, stdout, stderr = ssh.exec_command("cat /etc/pam.d/password-auth | grep auth | grep required | grep -E 'pam_faillock.so|pam_tally2.so' | grep -o 'deny=[0-9]*' | awk -F= '{print $2}'")
        ac03.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("cat /etc/pam.d/password-auth | grep auth | grep required | grep -E 'pam_faillock.so|pam_tally2.so' | grep -o 'unlock_time=[0-9]*' | awk -F= '{print $2}'")
        ac03.append(stdout.read().decode().strip())

        if ac03[0] == "" or int(ac03[0]) > 5:
            AC03[4] = "취약"
            AC03[3] = "deny : "+ac03[0]+", unlock_time : "+ac03[1]
        elif int(ac03[0]) > 5 or int(ac03[1]) > 1800:
            AC03[4] = "취약"
            AC03[3] = "deny : "+ac03[0]+", unlock_time : "+ac03[1]
        else:
            AC03[4] = "양호"
            AC03[3] = "deny : "+ac03[0]+", unlock_time : "+ac03[1]

        nsheet.append(AC03)
        if AC03[4] == "취약":
            nsheet['E7'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E7'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04A ==================================================
        AC04A = ["AC04-A", "사용자 계정, 그룹 리스트 정보 파일에 대한 접근권한이 제한되어 있는가?", "passwd : -rw-r--r--. 1 root root\ngroup : -rw-r--r--. 1 root root", "", ""]

        ac04a = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/passwd")
        ac04a.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/passwd")
        ac04a.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/group")
        ac04a.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/group")
        ac04a.append(stdout.read().decode().strip())

        AC04A[3] = "/etc/passwd : "+ac04a[0]+" "+ac04a[1]+", "+"/etc/group : "+ac04a[2]+" "+ac04a[3]

        AC04A[4] = "취약" if int(ac04a[0]) > 644 or ac04a[1] != "root" or int(ac04a[2]) > 644 or ac04a[3] != "root" else "양호"

        nsheet.append(AC04A)
        if AC04A[4] == "취약":
            nsheet['E8'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E8'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04B ==================================================
        AC04B = ["AC04-B", "사용자 계정 암호 파일에 대한 접근 권한이 제한되어 있는가?", "shadow : -r--------. 1 root root", "", ""]

        ac04b = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/shadow")
        ac04b.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/shadow")
        ac04b.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("sudo passwd -S ian | awk '{print $2}'")
        ac04b.append(stdout.read().decode().strip())

        AC04B[3] = "/etc/shadow : "+ac04b[0]+" "+ac04b[1]

        AC04B[4] = "취약" if int(ac04b[0]) > 400 or ac04b[1] != "root" or 'P' not in ac04b[2] else "양호"

        nsheet.append(AC04B)
        if AC04B[4] == "취약":
            nsheet['E9'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E9'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04C ==================================================
        AC04C = ["AC04-C", "/etc/hosts 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/hosts : -rw-r--r--. 1 root root", "", ""]

        ac04c = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/hosts")
        ac04c.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/hosts")
        ac04c.append(stdout.read().decode().strip())

        AC04C[3] = "/etc/hosts : "+ac04c[0]+" "+ac04c[1]

        AC04C[4] = "취약" if int(ac04c[0]) > 644 or ac04c[1] != "root" else "양호"

        nsheet.append(AC04C)
        if AC04C[4] == "취약":
            nsheet['E10'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E10'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04D ==================================================
        AC04D = ["AC04-D", "/etc/services 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/services : -rw-r--r--. 1 root root", "", ""]

        ac04d = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/services")
        ac04d.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/services")
        ac04d.append(stdout.read().decode().strip())

        AC04D[3] = "/etc/services : "+ac04d[0]+" "+ac04d[1]

        AC04D[4] = "취약" if int(ac04d[0]) > 644 or ac04d[1] != "root" else "양호"

        nsheet.append(AC04D)
        if AC04D[4] == "취약":
            nsheet['E11'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E11'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04E ==================================================
        AC04E = ["AC04-E", "일반 사용자가 중요 명령어를 실행하지 못하도록 설정되었는가?", "/usr/bin/last : -rw-------. 1 root root", "", ""]

        ac04e = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /usr/bin/last")
        ac04e.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /usr/bin/last")
        ac04e.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("command -v ifconfig")
        ac04e.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /usr/sbin/ifconfig")
        ac04e.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /usr/sbin/ifconfig")
        ac04e.append(stdout.read().decode().strip())

        if not ac04e[2]:
            AC04E[3] = "/usr/bin/last : "+ac04e[0]+" "+ac04e[1]
            AC04E[4] = "취약" if int(ac04e[0]) > 700 or ac04e[1] != "root" else "양호"
        else:
            AC04E[3] = "/usr/bin/last : "+ac04e[0]+" "+ac04e[1]+"\n"+"/usr/sbin/ifconfig : "+ac04e[3]+" "+ac04e[4]+"\n"
            AC04E[4] = "취약" if int(ac04e[0]) > 700 or ac04e[1] != "root" or int(ac04e[3]) > 700 or ac04e[4] != "root" else "양호"

        nsheet.append(AC04E)
        if AC04E[4] == "취약":
            nsheet['E12'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E12'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04F ==================================================
        AC04F = ["AC04-F", "사용자 환경 설정 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/profile : -rwxr-xr-x. 1 root root", "", ""]

        ac04f = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/profile")
        ac04f.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/profile")
        ac04f.append(stdout.read().decode().strip())

        AC04F[3] = "/etc/profile : "+ac04f[0]+" "+ac04f[1]+"\n"

        AC04F[4] = "취약" if int(ac04f[0]) > 755 or ac04f[1] != "root" else "양호"

        nsheet.append(AC04F)
        if AC04F[4] == "취약":
            nsheet['E13'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E13'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04G ==================================================
        AC04G = ["AC04-G", "네트워크 서비스 설정 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/xinetd.conf : -rw-------. 1 root root", "", ""]

        ac04g = []

        stdin, stdout, stderr = ssh.exec_command("ls -al /etc/xinetd.conf")
        ac04g.append(stdout.read().decode().strip())

        if len(ac04g[0]) == 0:
            AC04G[3] = "/etc/xinetd.conf 파일 없음"
            AC04G[4] = "양호"
        else:
            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/xinetd.conf")
            ac04g.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/xinetd.conf")
            ac04g.append(stdout.read().decode().strip())

            AC04G[3] = "/etc/xinetd.conf : "+ac04f[0]+" "+ac04f[1]+"\n"

            AC04G[4] = "취약" if int(ac04g[0]) > 600 or ac04g[1] != "root" else "양호"

        nsheet.append(AC04G)
        if AC04G[4] == "취약":
            nsheet['E14'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E14'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04H ==================================================
        AC04H = ["AC04-H", "FTP 접근제어 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/vsftpd.ftpusers, /etc/vsftpd.user_list : -rw-------. 1 root root", "", ""]

        ac04h = []

        stdin, stdout, stderr = ssh.exec_command("sudo systemctl status vsftpd | grep Active: | awk -F' ' '{print $2}'")
        ac04h.append(stdout.read().decode().strip())

        if ac04h[0] == "" or ac04h[0] == "inactive":
            AC04H[3] = "vsftp 서비스 기동되어있지 않음"
            AC04H[4] = "양호"
        else:
            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/vsftpd/ftpusers")
            ac04h.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/vsftpd/ftpusers")
            ac04h.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/vsftpd/user_list")
            ac04h.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/vsftpd/user_list")
            ac04h.append(stdout.read().decode().strip())

            AC04H[3] = "/etc/vsftpd/ftpusers : "+ac04h[1]+" "+ac04h[2]+", "+"/etc/vsftpd/user_list : "+ac04h[3]+" "+ac04h[4]
            AC04H[4] = "취약" if int(ac04h[1]) > 600 or ac04h[2] != "root" or int(ac04h[3]) > 600 or ac04h[4] != "root" else "양호"

        nsheet.append(AC04H)
        if AC04H[4] == "취약":
            nsheet['E15'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E15'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04I ==================================================
        AC04I = ["AC04-I", "NFS 설정 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/exports : -rw-------. 1 root root", "", ""]

        ac04i = []

        stdin, stdout, stderr = ssh.exec_command("ls -al /etc/exports")
        ac04i.append(stdout.read().decode().strip())

        if len(ac04i[0]) == 0:
            AC04I[3] = "NFS 데몬 설치 되어있지 않음"
            AC04I[4] = "양호"
        else:
            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/exports")
            ac04i.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%A' /etc/exports")
            ac04i.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/exports")
            ac04i.append(stdout.read().decode().strip())

            AC04I[3] = "/etc/exports : "+ac04i[1]+" "+ac04i[3]
            AC04I[4] = "취약" if "w" in ac04i[1][4:] or ac04i[3] != "root" else "양호"

        nsheet.append(AC04I)
        if AC04I[4] == "취약":
            nsheet['E16'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E16'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04J ==================================================
        AC04J = ["AC04-J", "패스워드 규칙 설정 파이에 대한 접근 권한이 제한되어 있는가?", "/etc/pam.d/system-auth, password-auth : 타사용자 쓰기권한 없음, 소유자 root", "", ""]

        ac04j = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/pam.d/system-auth")
        ac04j.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%A' /etc/pam.d/system-auth")
        ac04j.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/pam.d/system-auth")
        ac04j.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/pam.d/password-auth")
        ac04j.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%A' /etc/pam.d/password-auth")
        ac04j.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/pam.d/password-auth")
        ac04j.append(stdout.read().decode().strip())

        AC04J[3] = "/etc/pam.d/system-auth : "+ac04j[0]+" "+ac04j[2]+", "+"/etc/pam.d/password-auth : "+ac04j[3]+" "+ac04j[5]
        AC04J[4] = "취약" if "w" in ac04j[1][4:] or ac04j[2] != "root" or "w" in ac04j[4][4:] or ac04j[5] != "root" else "양호"

        nsheet.append(AC04J)
        if AC04J[4] == "취약":
            nsheet['E17'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E17'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04K ==================================================
        AC04K = ["AC04-K", "cron 설정 파일에 대한 접근 권한이 제한되어 있는가?", "/etc/cron.allow, cron.deny : 타사용자 쓰기권한 없음, 소유자 root", "", ""]

        ac04k = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/cron.allow")
        ac04k.append(stdout.read().decode().strip())

        if len(ac04k[0]) != 0:
            stdin, stdout, stderr = ssh.exec_command("stat -c '%A' /etc/cron.allow")
            ac04k.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/cron.allow")
            ac04k.append(stdout.read().decode().strip())

            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/cron.deny")
            ac04k.append(stdout.read().decode().strip())
            if len(ac04k[3]) != 0:
                stdin, stdout, stderr = ssh.exec_command("stat -c '%A' /etc/cron.deny")
                ac04k.append(stdout.read().decode().strip())

                stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/cron.deny")
                ac04k.append(stdout.read().decode().strip())

                AC04K[3] = "/etc/cron.allow : "+ac04k[0]+" "+ac04k[2]+", "+"/etc/cron.deny : "+ac04k[3]+" "+ac04k[5]
                AC04K[4] = "취약" if "w" in ac04k[1][4:] or ac04k[2] != "root" or "w" in ac04k[4][4:] or ac04k[5] != "root" else "양호"
            else:
                AC04K[3] = "/etc/cron.allow : "+ac04k[0]+" "+ac04k[2]+" /etc/cron.deny 파일 없음"
                AC04K[4] = "취약" if "w" in ac04k[1][4:] or ac04k[2] != "root" else "양호"
        else:
            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /etc/cron.deny")
            ac04k.append(stdout.read().decode().strip())
            if len(ac04k[1]) != 0:
                stdin, stdout, stderr = ssh.exec_command("stat -c '%A' /etc/cron.deny")
                ac04k.append(stdout.read().decode().strip())

                stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /etc/cron.deny")
                ac04k.append(stdout.read().decode().strip())

                AC04K[3] = "/etc/cron.deny : "+ac04k[1]+" "+ac04k[3]+" /etc/cron.allow 파일 없음"
                AC04K[4] = "취약" if "w" in ac04k[2][4:] or ac04k[3] != "root" else "양호"
            else:
                AC04K[3] = "/etc/cron.deny, cron.allow 파일 없음"
                AC04K[4] = "양호"

        nsheet.append(AC04K)
        if AC04K[4] == "취약":
            nsheet['E18'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E18'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-04L ==================================================
        AC04L = ["AC04-L", "주요 백업 파일 접근 권한이 제한되어 있는가?", "passwd, services, hosts, wtmp, btmp, sulog 백업 파일 유무 및 권한 확인", "", ""]

        stdin, stdout, stderr = ssh.exec_command("ls -al /etc/passwd* /etc/services* /etc/hosts* /var/log/wtmp* /var/log/btmp* /var/log/secure* | awk -F' ' '{print $9}'")
        file_lst = []
        for line in stdout:
            file_name = line.rstrip('\n')
            if file_name not in ["/etc/passwd", "/etc/services", "/etc/hosts", "/var/log/wtmp", "/var/log/btmp", "/var/log/secure", "/etc/hosts.allow", "/etc/hosts.deny"]:
                file_lst.append(file_name)

        AC04L[4] = "양호"
        for file in file_lst:
            stdin, stdout, stderr = ssh.exec_command("stat -c '%a' "+file)
            perm = stdout.read().decode().strip()

            stdin, stdout, stderr = ssh.exec_command("stat -c '%U' "+file)
            own = stdout.read().decode().strip()
            if file == "/etc/passwd-":
                if int(perm) > 644 or own != "root":
                    AC04L[4] = "취약"
            else:
                if int(perm) != 600 or own != "root":
                    AC04L[4] = "취약"
            AC04L[3] += file+" : "+perm+" "+own+" "

        nsheet.append(AC04L)
        if AC04L[4] == "취약":
            nsheet['E19'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E19'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-05A ==================================================
        AC05A = ["AC05-A", "root 계정의 UMASK 설정은 적절한가?", "UMASK : 022, 027", "", ""]

        ac05a = []

        stdin, stdout, stderr = ssh.exec_command("sudo su - root -c 'umask'")
        ac05a.append(stdout.read().decode().rstrip('\n'))

        AC05A[3] = "UMASK : "+ac05a[0]
        AC05A[4] = "취약" if int(ac05a[0]) > 27 else "양호"

        nsheet.append(AC05A)
        if AC05A[4] == "취약":
            nsheet['E20'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E20'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-05B ==================================================
        AC05B = ["AC05-B", "일반 계정의 UMASK 설정은 적절한가?", "UMASK : 022, 027", "", ""]

        ac05b = []

        AC05B[4] = "양호"
        for user in user_lst:
            stdin, stdout, stderr = ssh.exec_command("sudo su - "+user+" -c 'umask'")
            output = stdout.read().decode().strip()
            error = stdout.read().decode().strip()

            #ret = int(stdout.read().decode().strip())
            #if ret > 27:
            try:
                ret = int(output)
                if ret > 27:
                    ac05b.append(user+" : "+str(ret))
                    AC05B[4] = "취약"
                else:
                    ac05b.append(user+" : "+str(ret))
            except ValueError:
                ac05b.append(user + " : umask 조회 불가 (" + (output or error) + ")")

        AC05B[3] = ", ".join(ac05b)

        nsheet.append(AC05B)
        if AC05B[4] == "취약":
            nsheet['E21'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E21'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-05C ==================================================
        AC05C = ["AC05-C", "root 계정의 원격접속은 제한되어 있는가?", "/etc/ssh/sshd_config : PermitRootLogin no", "", ""]

        ac05c = []

        stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/ssh/sshd_config | grep 'PermitRootLogin' | grep -v '#' | awk '{print $2}'")
        ac05c.append(stdout.read().decode().strip())
        
        AC05C[3] = "PermitRootLogin : "+ac05c[0]
        AC05C[4] = "양호" if ac05c[0] == "no" else "취약"

        nsheet.append(AC05C)
        if AC05C[4] == "취약":
            nsheet['E22'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E22'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== AC-06 ==================================================
        AC06 = ["AC06", "Session Timeout을 적용하고 있는가?", "TMOUT : 1800", "", ""]

        ac06 = []

        stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/profile | grep TMOUT | grep -v '#' | awk -F= '{print $2}'")
        ac06.append(stdout.read().decode().strip())

        AC06[4] = "양호"
        if len(ac06[0]) == 0:
            AC06[4] = "취약"
            ac06[0] = "TMOUT 설정 없음"
            AC06[3] = "/etc/profile TMOUT : "+ac06[0]
        elif int(ac06[0]) > 1800:
            AC06[4] = "취약"
            AC06[3] = "/etc/profile TMOUT : "+ac06[0]
        else:
            for user in user_lst:
                stdin, stdout, stderr = ssh.exec_command("cat /home/"+user+"/.bash_profile | grep TMOUT | grep -v '#' | awk -F= '{print $2}'")
                ret = stdout.read().decode().strip()
                if len(ret) == 0:
                    ac06.append(user+" : 0")
                elif int(ret) > 1800:
                    AC06[4] = "취약"
                    ac06.append(user+" TMOUT : "+ret)
                else:
                    u01.append(user+" TMOUT : "+ret)

            AC06[3] = "/etc/profile TMOUT : "+ac06[0]+", .bash_profile TMOUT : ".join(ac06[1:])

        nsheet.append(AC06)
        if AC06[4] == "취약":
            nsheet['E23'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E23'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== SS-01 ==================================================
        SS01 = ["SS01", "NFS 사용 시 안전한 설정을 적용 하였는가?", "공유 디렉토리 : root, 사용자 홈디렉토리 공유 금지, 권한 : everyone 공유 금지", "", ""]

        ss01 = []

        stdin, stdout, stderr = ssh.exec_command("sudo systemctl status nfs-server | grep Active: | awk -F' ' '{print $2}'")
        ss01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("sudo netstat -antp | grep 2049 | grep ESTABLISHED | wc -l")
        ss01.append(stdout.read().decode().strip())

        if ss01[0] == "Active" and int(ss01[1]) > 0:
            stdin, stdout, stderr = ssh.exec_command("cat /etc/exports")
            ss01.append(stdout.read().decode().strip())
            if "/home" in sso1[2] or "/root" in ss01[2] or "*" in ss01[2]:
                SS01[4] = "취약"
                SS01[3] = "NFS-Server 서비스: "+ss01[0]+", NFS 연결 수: "+ss01[1]+", /etc/exports 설정: "+ss01[2]
            else:
                SS01[4] = "양호"
                SS01[3] = "NFS-Server 서비스: "+ss01[0]+", NFS 연결 수: "+ss01[1]+", /etc/exports 설정: "+ss01[2]
        elif ss01[0] == "Active":
            stdin, stdout, stderr = ssh.exec_command("cat /etc/exports")
            ss01.append(stdout.read().decode().strip())
            if "/home" in u01[2] or "/root" in u01[2] or "*" in u01[2]:
                SS01[4] = "취약"
                SS01[3] = "NFS-Server 서비스: "+ss01[0]+", NFS 연결 수: "+ss01[1]+", /etc/exports 설정: "+ss01[2]
            else:
                SS01[4] = "양호"
                SS01[3] = "NFS-Server 서비스: "+ss01[0]+", NFS 연결 수: "+ss01[1]+", /etc/exports 설정: "+ss01[2]
        elif int(ss01[1]) > 0:
            stdin, stdout, stderr = ssh.exec_command("cat /etc/fstab | grep nfs")
            ss01.append(stdout.read().decode().strip())
            if "/root" in ss01[2] or "/home" in ss01[2]:
                SS01[4] = "취약"
                SS01[3] = "NFS Client 동작중, NFS 연결 수: "+ss01[1]+", /etc/fstab 설정: "+ss01[2]
            else:
                SS01[4] = "양호"
                SS01[3] = "NFS Client 동작중, NFS 연결 수: "+ss01[1]+", /etc/fstab 설정: "+ss01[2]
        else:
            SS01[4] = "양호"
            SS01[3] = "NFS Server 및 Client로 동작안함"

        nsheet.append(SS01)
        if SS01[4] == "취약":
            nsheet['E24'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E24'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== SS-02A ==================================================
        SS02A = ["SS02-A", "원격에서 인증없이 접속하여 명령 실행이 가능한 서비스 사용을 중지하였는가?", "r계열 서비스 중지 확인", "", ""]

        ss02a = []

        stdin, stdout, stderr = ssh.exec_command("rpm -qa | grep -E 'rsh|rlogin|rexec' | wc -l")
        ss02a.append(stdout.read().decode().strip())
        
        SS02A[3] = "r계열 서비스 설치수: "+ss02a[0]
        SS02A[4] = "양호" if int(ss02a[0]) == 0 else "취약"

        nsheet.append(SS02A)
        if SS02A[4] == "취약":
            nsheet['E25'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E25'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== SS-02B ==================================================
        SS02B = ["SS02-B", "불필요한 서비스는 중지 또는 제거하였는가?", "불필요 서비스 중지 확인", "", ""]

        ss02b = []
        stdin, stdout, stderr = ssh.exec_command("sudo systemctl status time")
        ss02b.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("sudo systemctl list-units --type=service | grep -E 'bootp|chargen|cmsd|daytime|discard|echo|finger|netstat|rusersd|sprayed|systat|tftp|ttdbserverd|uucp'")
        for line in stdout:
            ss02b.append(line.rstrip('\n'))

        if len(ss02b) <= 1 and len(ss02b[0]) == 0:
            SS02B[4] = "양호"
            SS02B[3] = "불필요한 서비스 없음"
        else:
            SS02B[4] = "취약"
            SS02B[3] = "불필요 서비스 리스트 : "+", ".join(ss02b)

        nsheet.append(SS02B)
        if SS02B[4] == "취약":
            nsheet['E26'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E26'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== SS-03A ==================================================
        SS03A = ["SS03-A", "SNMP Default Community String이 변경되었는가?", "Community string : public, private 등 default 값 사용 금지 ", "", ""]

        ss03a = []

        stdin, stdout, stderr = ssh.exec_command("sudo systemctl status snmpd | grep Active: | awk -F' ' '{print $2}'")
        ss03a.append(stdout.read().decode().strip())

        if ss03a[0] == "inactive":
            SS03A[4] = "양호"
            SS03A[3] = "SNMP 서비스 동작 안함"
        else:
            stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/snmp/snmpd.conf | grep -v '#' | grep com2sec")
            ss03a.append(stdout.read().decode().strip())
            if "public" in ss03a[1]:
                SS03A[4] = "취약"
                SS03A[3] = "SNMP 서비스 : "+ss03a[0]+", community 설정 : "+ss03a[1]
            else:
                SS03A[4] = "양호"
                SS03A[3] = "SNMP 서비스 : "+ss03a[0]+", community 설정 : "+ss03a[1]

        nsheet.append(SS03A)
        if SS03A[4] == "취약":
            nsheet['E27'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E27'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== SS-03B ==================================================
        SS03B = ["SS03-B", "Anonymouse FTP가 중지되어 있는가?", "/etc/passwd, shadow : ftp, anonymous 계정 없음 확인, anonymous_enable : No", "", ""]

        ss03b = []

        stdin, stdout, stderr = ssh.exec_command("sudo systemctl status vsftpd | grep Active: | awk -F' ' '{print $2}'")
        ss03b.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/passwd | grep -E 'ftp|anonymous' | awk -F: '{print $1}'")
        for line in stdout:
            ss03b.append(line.rstrip('\n'))

        stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/shadow | grep -E 'ftp|anonymous' | awk -F: '{print $1}'")
        for line in stdout:
            ss03b.append(line.rstrip('\n'))

        if ss03b[0] == "Active":
            stdin, stdout, stderr = ssh.exec_command("sudo cat /etc/vsftpd/vsftpd.conf | grep anonymous | grep -v '#' | awk -F= '{print $2}'")
            ret = stdout.read().decode().strip()
            if ret == "YES" or "ftp" in ss03b or "anonymous" in ss03b:
                SS03B[4] = "취약"
                if len(ss03b[1:]) == 0:
                    SS03B[3] = "passwd, shadow에 ftp, anonymous계정 없음, anonymous_enable : "+ret
                else:
                    SS03B[3] = "passwd, shadow : "+", ".join(ss03b[1:])+", anonymous_enable : "+ret
            else:
                SS03B[4] = "양호"
                SS03B[3] = "passwd, shadow에 ftp, anonymous계정 없음, anonymous_enable : "+ret
        else:
            if "ftp" in ss03b or "anonymous" in ss03b:
                SS03B[4] = "취약"
                SS03B[3] = "passwd, shadow : "+", ".join(ss03b[1:])+", vsftpd 서비스 없음"
            else:
                SS03B[4] = "양호"
                SS03B[3] = "passwd, shadow에 ftp, anonymous계정 없음, vsftpd 서비스 없음"

        nsheet.append(SS03B)
        if SS03B[4] == "취약":
            nsheet['E28'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E28'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== SS-04 ==================================================
        SS04 = ["SS04", "root 계정의 Path 설정이 안전하게 적용되어 있는가?", "Path 환경 변수 : 현재 디렉토리(.) 미포함", "", ""]

        ss04 = []

        stdin, stdout, stderr = ssh.exec_command("sudo su - root -c 'env | grep PATH'")
        ss04.append(stdout.read().decode().strip())

        SS04[3] = "PATH 환경 변수 : "+ss04[0]
        SS04[4] = "취약" if ".:" in ss04[0] else "양호"

        nsheet.append(SS04)
        if SS04[4] == "취약":
            nsheet['E29'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E29'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# ================================================== LP-01 ==================================================
        LP01 = ["LP01", "서버 접속 로그에 대한 접근 권한을 제한 하는가?", "/var/log/wtmp, btmp, secure : root 600, messages : root 644", "", ""]

        lp01 = []

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /var/log/wtmp")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /var/log/wtmp")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /var/log/btmp")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /var/log/btmp")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /var/log/secure")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /var/log/secure")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%a' /var/log/messages")
        lp01.append(stdout.read().decode().strip())

        stdin, stdout, stderr = ssh.exec_command("stat -c '%U' /var/log/messages")
        lp01.append(stdout.read().decode().strip())

        LP01[3] = "/var/log/wtmp, btmp, secure : "+lp01[0]+" "+lp01[1]+", "+lp01[2]+" "+lp01[3]+", "+lp01[3]+" "+lp01[4]+"/var/log/messages : "+lp01[5]+" "+lp01[6]
        LP01[4] = "취약" if lp01[0] != "600" or lp01[1] != "root" or lp01[2] != "600" or lp01[3] != "root" or lp01[4] != "600" or lp01[5] != "root" or lp01[6] != "644" or lp01[7] != "root" else "양호"

        nsheet.append(LP01)
        if LP01[4] == "취약":
            nsheet['E30'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        else:
            nsheet['E30'].fill = PatternFill(fill_type="solid", start_color="00ff00")
# =================================================== END ==================================================

        for i in range(1, 6):
            for j in range(1, 31):
                nsheet.cell(row=j, column=i).border = thin_border

    except paramiko.ssh_exception.NoValidConnectionsError as 접속불가:
        print("\n["+svr[4]+"]"+" [접속불가] IP 및 PORT 확인")
    
    except paramiko.ssh_exception.AuthenticationException as 접속불가:
        print("\n["+svr[4]+"]"+" [접속불가] 사용자 이름 또는 패스워드 확인")

    except paramiko.ssh_exception.PasswordRequiredException as 접속불가:
        print("\n["+svr[4]+"]"+" [접속불가] Password Required")

    except paramiko.ssh_exception.SSHException as 접속불가:
        print("\n["+svr[4]+"]"+" [접속불가] SSHException")
    finally:
        ssh.close()

#for idx, svr in enumerate(result):
for idx, svr in tqdm(enumerate(svr_list), total = len(svr_list), desc = '진행률', ncols = 70, ascii = ' =', leave = True,):
    if svr[5] == 'redhat':
        rsec(idx, svr)
    else:
        usec(idx, svr)

wb.save(filename="Unix_sec_chk.xlsx")
